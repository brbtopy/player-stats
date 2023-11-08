from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from win32com import client
import os
import shutil
import requests
import json
from stat import S_IREAD, S_IRGRP, S_IROTH
from stat import S_IWUSR # Need to add this import to the ones above

try:
  workbook1 = load_workbook(filename="Renaissance FC Player Inventory.xlsx")
  workbook2 = load_workbook(filename="Indv. Player Profile.xlsx")

  filename = "Indv. Player Profile.xlsx"
  os.chmod(filename, S_IWUSR|S_IREAD) # This makes the file read/write for the owner

  General_Info_sheet = workbook1["General Info"]
  Player_Discipline_sheet = workbook1["Player Discipline"]
  Goal_Involvement_sheet = workbook1["Goal Involvement"]
  Indv_sheet = workbook2.active

  #GENERATING PLAYER ENTRY NUMBER TO UPDATE
  valid = False
  while not valid:
      try:
          name_cell = input("Give the index of the player whose individual profile is wanted: ")
          default_numbering = int(name_cell)
          name_cell_conc = "A" + str(default_numbering)
          valid = True
      except ValueError:
          print("[+] Incorrect data type [+]\n[+] Please type the correct entry position [+]")

  name_cell_list = list(name_cell_conc)
  first = ""
  second = ""
  third = ""
  name_cell_number = ""

  if len(name_cell_list) == 2:
      first = name_cell_list[1]
      name_cell_number = str(first)

  elif len(name_cell_list) == 3:
      first = name_cell_list[1]
      second = name_cell_list[2]
      name_cell_number = str(first) + str(second)

  elif len(name_cell_list) == 4:
      first = name_cell_list[1]
      second = name_cell_list[2]
      third = name_cell_list[3]
      name_cell_number = str(first) + str(second) + str(third)
      
  player_name = General_Info_sheet["A" + name_cell_number].value

  #APPENDING GENERAL INFO DATA TO INDIVIDUAL PLAYER PROFILES
  Indv_sheet["D4"].value = General_Info_sheet["A" + name_cell_number].value
  Indv_sheet["D5"].value = General_Info_sheet["G" + name_cell_number].value
  Indv_sheet["D6"].value = General_Info_sheet["H" + name_cell_number].value
  Indv_sheet["D7"].value = General_Info_sheet["D" + name_cell_number].value
  Indv_sheet["D8"].value = General_Info_sheet["E" + name_cell_number].value
  Indv_sheet["D9"].value = General_Info_sheet["C" + name_cell_number].value
  Indv_sheet["D9"].alignment = Alignment(horizontal='left', vertical='bottom') #align text in cell to the left
  Indv_sheet["D10"].value = General_Info_sheet["F" + name_cell_number].value
  Indv_sheet["D11"].value = General_Info_sheet["B" + name_cell_number].value
  Indv_sheet["D12"].value = General_Info_sheet["Q" + name_cell_number].value
  Indv_sheet["D13"].value = General_Info_sheet["L" + name_cell_number].value
  Indv_sheet["D14"].value = General_Info_sheet["M" + name_cell_number].value
  Indv_sheet["D15"].value = General_Info_sheet["K" + name_cell_number].value

  #APPENDING GOAL INVOLVEMENT STATS TO INDIVIDUAL PLAYER PROFILES
  Indv_sheet["D19"].value = Goal_Involvement_sheet["C" + name_cell_number].value
  Indv_sheet["E19"].value = Goal_Involvement_sheet["D" + name_cell_number].value
  Indv_sheet["F19"].value = Goal_Involvement_sheet["E" + name_cell_number].value
  Indv_sheet["G19"].value = Goal_Involvement_sheet["F" + name_cell_number].value

  #APPENDING OTHER COMPETITION STATS TO INDIVIDUAL PLAYER PROFILES
  Indv_sheet["D23"].value = Player_Discipline_sheet["C" + name_cell_number].value
  Indv_sheet["E23"].value = Player_Discipline_sheet["D" + name_cell_number].value

  #APPENDING EDUCATION STATS TO INDIVIDUAL PLAYER PROFILES
  Indv_sheet["A29"].value = General_Info_sheet["R" + name_cell_number].value
  Indv_sheet["D29"]. value = General_Info_sheet["S" + name_cell_number].value

  #APPENDING PERSONAL INFO TO INDIVIDUAL PLAYER PROFILES
  Indv_sheet["B32"].value = General_Info_sheet["I" + name_cell_number].value
  Indv_sheet["B33"].value = General_Info_sheet["J" + name_cell_number].value
  Indv_sheet["B34"].value = General_Info_sheet["O" + name_cell_number].value
  Indv_sheet["B35"].value = General_Info_sheet["P" + name_cell_number].value
  Indv_sheet["B35"].font = Font(name='Calibri', size=12) #change font name and size to calibri and size 12

  workbook1.close()
  workbook2.save(filename="Indv. Player Profile.xlsx")
  workbook2.close()

  #CONVERT EXCEL TO PDF
  print("[+] Converting Excel file to PDF format ... [+]")
  def excl2pdf(file_location):
    app = client.DispatchEx("Excel.Application")
    app.Interactive = False
    app.Visible = False

    workbook = app.Workbooks.open(file_location)
    output = os.path.splitext(file_location)[0]
    
    workbook.ActiveSheet.ExportAsFixedFormat(0, output)
    workbook.Close()

  excl2pdf("C:\\Users\\samas\\Desktop\\proj\\rfc\\rfc-player-stats\\Indv. Player Profile.xlsx")
  print("[+] Done converting Excel to PDF [+]\n")

  os.chmod(filename, S_IREAD|S_IRGRP|S_IROTH)

  #ADD WATERMARK TO PDF
  print("[+] Adding watermark to PDF ... [+]")
  instructions = {
    'parts': [
      {
        'file': 'document'
      }
    ],
    'actions': [
      {
        'type': 'watermark',
        'image': 'logo',
        'width': '70%',
        "opacity": 0.1
      }
    ]
  }

  output_file = f"{player_name}'s Indv. Player Profile.pdf"

  response = requests.request(
    'POST',
    'https://api.pspdfkit.com/build',
    headers = {
      'Authorization': 'Bearer pdf_live_09Jw3fmsDx53azf2r0TEjwhxqlqriaVRV3TGCHtl539'
    },
    files = {
      'document': open('Indv. Player Profile.pdf', 'rb'),
      'logo': open('pictures\\logo.png', 'rb')
    },
    data = {
      'instructions': json.dumps(instructions)
    },
    stream = True
  )

  if response.ok:
    with open(output_file, 'wb') as fd:
      for chunk in response.iter_content(chunk_size=8096):
        fd.write(chunk)
  else:
    print(response.text)
    exit()

  os.remove("Indv. Player Profile.pdf")
  print("[+] Watermark added. Exiting program [+]")

  shutil.move(output_file, f"indv player profiles\\{output_file}")

except KeyboardInterrupt:
  print("\n[+] Programme was interrupted by the user. Exiting ... [+]")
  os.chmod(filename, S_IREAD|S_IRGRP|S_IROTH)
  pass