from openpyxl import load_workbook
import os
from stat import S_IREAD, S_IRGRP, S_IROTH
from stat import S_IWUSR # Need to add this import to the ones above

workbook = load_workbook(filename="Renaissance FC Player Inventory.xlsx")
General_Info_sheet = workbook["General Info"]
Player_Discipline_sheet = workbook["Player Discipline"]
Goal_Involvement_sheet = workbook["Goal Involvement"]

filename = "Renaissance FC Player Inventory.xlsx"
os.chmod(filename, S_IWUSR|S_IREAD) # This makes the file read/write for the owner

print("[+] This programme updates the penalty statistics of a player [+]\n")
print("[+] Type and enter 'end' or 'stop' to end programme [+]")
player_name = input("[+] Which player attempted a penalty? ")

while player_name != "end":
    for rowNum in range(2, int(General_Info_sheet.max_row)+1):
        name_cell = General_Info_sheet.cell(row=rowNum, column=1).value
        if name_cell.title() == player_name.title():
            Goal_Involvement_sheet["E" + str(rowNum)].value += 1
            check_penalty = input(f"Did {player_name.title()} score the penalty? Enter 'yes' or 'no': ")
            if check_penalty == "yes":
                Goal_Involvement_sheet["F" + str(rowNum)].value += 1
            else:
                print(f"[+] {player_name.title()} didn't make the penalty [+]")
            print("[+] Penalty statistic has been updated successfully [+]\n")

    workbook.save(filename="Renaissance FC Player Inventory.xlsx")
    player_name = input("[+] Which player attempted a penalty? ")
      

workbook.save(filename="Renaissance FC Player Inventory.xlsx")
workbook.close()

os.chmod(filename, S_IREAD|S_IRGRP|S_IROTH)